"""
Write reconciled results back to a copy of the input Excel.

Uses openpyxl to preserve the original formatting (column widths,
colors, filters). The original file is NEVER modified; a new file
with "_filled" suffix is created.
"""
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# Color codes for visual highlighting
COLOR_AUTO_OK = "C6EFCE"       # light green
COLOR_LOW_CONFIDENCE = "FFEB9C"  # light yellow
COLOR_PRODUCT_GONE = "F4CCCC"   # light red


def make_output_path(input_path: str) -> str:
    """Generate the output path: input.xlsx -> input_filled.xlsx"""
    p = Path(input_path)
    return str(p.with_stem(p.stem + "_filled"))


def write_reconciliation(
    input_path: str,
    output_path: str,
    updates: list,
) -> None:
    """
    Copy input Excel to output path, then apply updates.

    Args:
        input_path: original Excel file path.
        output_path: new file to be created.
        updates: list of dicts, each with:
            {
                "row_index": int,        # 1-indexed Excel row (so header is row 1)
                "remark_l1": str | None,
                "remark_l2": str | None,
                "decision": str,         # for color coding
            }
    """
    # 1. Copy the original file so we never modify it in place
    shutil.copyfile(input_path, output_path)

    # 2. Open the copy and update
    wb = load_workbook(output_path)
    ws = wb.active

    # 3. Find the column indices for Remarks L1 and Remarks L2
    header_row = 1
    col_l1 = None
    col_l2 = None
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        if header_value == "Remarks L1":
            col_l1 = col_idx
        elif header_value == "Remarks L2 (details)":
            col_l2 = col_idx

    if col_l1 is None or col_l2 is None:
        raise ValueError(
            f"Could not find 'Remarks L1' or 'Remarks L2 (details)' columns. "
            f"Found col_l1={col_l1}, col_l2={col_l2}"
        )

    # 4. Apply each update
    for u in updates:
        row = u["row_index"]
        ws.cell(row=row, column=col_l1).value = u["remark_l1"]
        ws.cell(row=row, column=col_l2).value = u["remark_l2"]

        # Color-code based on decision
        fill_color = _decision_to_color(u["decision"])
        if fill_color:
            fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type="solid",
            )
            ws.cell(row=row, column=col_l1).fill = fill
            ws.cell(row=row, column=col_l2).fill = fill

    # 5. Save
    wb.save(output_path)


def _decision_to_color(decision: str) -> str:
    """Map a reconciliation decision label to a cell fill color."""
    if decision.startswith("AUTO_PRODUCT_GONE") or \
       decision.startswith("AUTO_CONFIRMED_MISSING"):
        return COLOR_PRODUCT_GONE
    if decision.startswith("AUTO_"):
        return COLOR_AUTO_OK
    if decision.startswith("LOW_CONFIDENCE"):
        return COLOR_LOW_CONFIDENCE
    return None